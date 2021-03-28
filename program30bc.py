from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
#Přidala jsem si ještě dny
dny = ["PONDĚLÍ", "ÚTERÝ", "STŘEDA", "ČTVRTEK", "PÁTEK"]

cervena = PatternFill("solid", fgColor="00FF0000")
zluta = PatternFill("solid", fgColor="00FFFF00")

radek = 1
for den in rozvrh_hodin:
  bunka = ws1.cell(radek, 1)
  bunka.value = dny[radek-1]
  bunka.fill = cervena
  sloupec = 2
  for predmet in den:
    bunka = ws1.cell(radek, sloupec)
    bunka.value = predmet
    if bunka.value == "Oběd":
      bunka.fill = zluta
    sloupec += 1
  radek += 1

wb.save(filename="C:/Users/Zuzana/Documents/Python/rozvrh_hodin.xlsx")